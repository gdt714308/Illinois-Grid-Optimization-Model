import openpyxl
import pandas as pd
from storage_model import load_inputs

def run_transition_model(gen_data, policy, demand_df):

    wb = openpyxl.load_workbook('illinois_grid_model.xlsx')
    ws = wb['Inputs_Workforce']

    plants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plants.append({
                'name':               row[0],
                'type':               row[1],
                'workers':            row[2],
                'county':             row[3],
                'retirement_year':    row[4],
                'retrain_cost':       row[5],
            })
    plants_df = pd.DataFrame(plants)

    # Repurposing logic:
    # Coal plants -> nuclear or battery depending on size
    # Gas plants  -> battery storage
    # Threshold: plants with 400+ workers assumed large enough for nuclear conversion
    def assign_pathway(row):
        if row['type'] == 'Coal':
            if row['workers'] >= 400:
                return 'Nuclear_Conversion'
            else:
                return 'Battery_Storage'
        elif row['type'] == 'Natural_Gas':
            return 'Battery_Storage'
        return 'Decommission'

    # Capital cost by pathway (USD billions)
    # Nuclear conversion: DOE estimate $4-9B; use $6B mid for large plants
    # Battery storage:    $125-334/KWh for 4hr BESS; assume 500 MWh site = ~$115M
    def assign_capex(row):
        if row['pathway'] == 'Nuclear_Conversion':
            return 6.0   # $6B
        elif row['pathway'] == 'Battery_Storage':
            return 0.115  # $115M
        return 0.010      # $10M decommission

    # Build timeline: for each plant, construction starts 2 years before retirement
    # Nuclear takes 10 years to build so starts much earlier
    def assign_construction_start(row):
        if row['pathway'] == 'Nuclear_Conversion':
            return row['retirement_year'] - 10
        elif row['pathway'] == 'Battery_Storage':
            return row['retirement_year'] - 2
        return row['retirement_year'] - 1

    plants_df['pathway']            = plants_df.apply(assign_pathway, axis=1)
    plants_df['capex_billions']     = plants_df.apply(assign_capex, axis=1)
    plants_df['construction_start'] = plants_df.apply(assign_construction_start, axis=1)
    plants_df['total_retrain_cost'] = plants_df['workers'] * plants_df['retrain_cost'] / 1e6  # $M

    # Year by year capital spend (spread evenly across construction window)
    years = list(range(2025, 2046))
    annual_capex = {y: 0.0 for y in years}
    annual_workers_retrained = {y: 0 for y in years}

    for _, row in plants_df.iterrows():
        start = max(row['construction_start'], 2025)
        end   = row['retirement_year']
        span  = max(end - start, 1)
        annual_spend = row['capex_billions'] / span

        for y in range(start, end + 1):
            if y in annual_capex:
                annual_capex[y] += annual_spend

        # Workers retrained in retirement year
        if row['retirement_year'] in annual_workers_retrained:
            annual_workers_retrained[row['retirement_year']] += row['workers']

    # Build summary table
    timeline_rows = []
    cumulative_capex = 0
    cumulative_workers = 0
    for y in years:
        cumulative_capex    += annual_capex[y]
        cumulative_workers  += annual_workers_retrained[y]
        timeline_rows.append({
            'year':                  y,
            'annual_capex_billions': round(annual_capex[y], 3),
            'cumulative_capex_B':    round(cumulative_capex, 3),
            'workers_retrained':     annual_workers_retrained[y],
            'cumulative_workers':    cumulative_workers,
        })
    timeline_df = pd.DataFrame(timeline_rows)

    # Print plant-level summary
    print("\nPLANT TRANSITION PLAN")
    print(f"{'Plant':<25} {'Type':<12} {'Pathway':<22} {'Retire':<8} {'Build Start':<12} {'CapEx $B':<10} {'Workers'}")
    print("-" * 105)
    for _, row in plants_df.iterrows():
        print(f"{row['name']:<25} {row['type']:<12} {row['pathway']:<22} {row['retirement_year']:<8} {row['construction_start']:<12} {row['capex_billions']:<10.3f} {row['workers']}")

    # Print year-by-year summary (key years only)
    print(f"\nYEAR-BY-YEAR CAPITAL SPEND AND WORKFORCE")
    print(f"{'Year':<8} {'Annual CapEx $B':<18} {'Cumulative $B':<16} {'Workers Retrained':<20} {'Cumulative Workers'}")
    print("-" * 75)
    key_years = [2025, 2026, 2027, 2028, 2029, 2030, 2033, 2035, 2037, 2040, 2042, 2044, 2045]
    for _, row in timeline_df[timeline_df['year'].isin(key_years)].iterrows():
        print(f"{int(row['year']):<8} {row['annual_capex_billions']:<18.3f} {row['cumulative_capex_B']:<16.3f} {int(row['workers_retrained']):<20} {int(row['cumulative_workers'])}")

    # Totals
    total_capex   = plants_df['capex_billions'].sum()
    total_workers = plants_df['workers'].sum()
    total_retrain = plants_df['total_retrain_cost'].sum()

    print(f"\n{'='*55}")
    print(f"TRANSITION TOTALS")
    print(f"  Plants transitioning:       {len(plants_df)}")
    print(f"  Total capital investment:   ${total_capex:.2f}B")
    print(f"  Total workers affected:     {total_workers:,}")
    print(f"  Total retraining cost:      ${total_retrain:.1f}M")
    print(f"  Nuclear conversions:        {len(plants_df[plants_df['pathway']=='Nuclear_Conversion'])}")
    print(f"  Battery conversions:        {len(plants_df[plants_df['pathway']=='Battery_Storage'])}")
    print(f"{'='*55}")

    return plants_df, timeline_df


if __name__ == '__main__':
    gen_data, policy, demand_df = load_inputs()
    plants_df, timeline_df = run_transition_model(gen_data, policy, demand_df)

def run_transition_model_v2(gen_data, policy, demand_df):
    """
    V2: Nuclear conversion only viable if construction_start >= 2025.
    Plants where 10-year build window doesn't fit get Battery_Storage instead.
    """
    wb = openpyxl.load_workbook('illinois_grid_model.xlsx')
    ws = wb['Inputs_Workforce']

    plants = []
    for row in ws.iter_rows(min_row=2, values_only=True):
        if row[0]:
            plants.append({
                'name':            row[0],
                'type':            row[1],
                'workers':         row[2],
                'county':          row[3],
                'retirement_year': row[4],
                'retrain_cost':    row[5],
            })
    plants_df = pd.DataFrame(plants)

    def assign_pathway_v2(row):
        if row['type'] == 'Coal':
            construction_start = row['retirement_year'] - 10
            if row['workers'] >= 400 and construction_start >= 2025:
                return 'Nuclear_Conversion'
            else:
                return 'Battery_Storage'
        elif row['type'] == 'Natural_Gas':
            return 'Battery_Storage'
        return 'Decommission'

    def assign_capex(row):
        if row['pathway'] == 'Nuclear_Conversion':
            return 6.0
        elif row['pathway'] == 'Battery_Storage':
            return 0.115
        return 0.010

    def assign_construction_start(row):
        if row['pathway'] == 'Nuclear_Conversion':
            return row['retirement_year'] - 10
        elif row['pathway'] == 'Battery_Storage':
            return row['retirement_year'] - 2
        return row['retirement_year'] - 1

    plants_df['pathway']            = plants_df.apply(assign_pathway_v2, axis=1)
    plants_df['capex_billions']     = plants_df.apply(assign_capex, axis=1)
    plants_df['construction_start'] = plants_df.apply(assign_construction_start, axis=1)
    plants_df['total_retrain_cost'] = plants_df['workers'] * plants_df['retrain_cost'] / 1e6

    years = list(range(2025, 2046))
    annual_capex = {y: 0.0 for y in years}
    annual_workers_retrained = {y: 0 for y in years}

    for _, row in plants_df.iterrows():
        start = max(row['construction_start'], 2025)
        end   = row['retirement_year']
        span  = max(end - start, 1)
        annual_spend = row['capex_billions'] / span
        for y in range(start, end + 1):
            if y in annual_capex:
                annual_capex[y] += annual_spend
        if row['retirement_year'] in annual_workers_retrained:
            annual_workers_retrained[row['retirement_year']] += row['workers']

    timeline_rows = []
    cumulative_capex = 0
    cumulative_workers = 0
    for y in years:
        cumulative_capex   += annual_capex[y]
        cumulative_workers += annual_workers_retrained[y]
        timeline_rows.append({
            'year':                  y,
            'annual_capex_billions': round(annual_capex[y], 3),
            'cumulative_capex_B':    round(cumulative_capex, 3),
            'workers_retrained':     annual_workers_retrained[y],
            'cumulative_workers':    cumulative_workers,
        })
    timeline_df = pd.DataFrame(timeline_rows)

    print("\nPLANT TRANSITION PLAN V2 (nuclear only where timeline is feasible)")
    print(f"{'Plant':<25} {'Type':<12} {'Pathway':<22} {'Retire':<8} {'Build Start':<12} {'CapEx $B':<10} {'Workers'}")
    print("-" * 105)
    for _, row in plants_df.iterrows():
        print(f"{row['name']:<25} {row['type']:<12} {row['pathway']:<22} {row['retirement_year']:<8} {row['construction_start']:<12} {row['capex_billions']:<10.3f} {row['workers']}")

    print(f"\nYEAR-BY-YEAR CAPITAL SPEND AND WORKFORCE")
    print(f"{'Year':<8} {'Annual CapEx $B':<18} {'Cumulative $B':<16} {'Workers Retrained':<20} {'Cumulative Workers'}")
    print("-" * 75)
    key_years = [2025,2026,2027,2028,2029,2030,2033,2035,2037,2040,2042,2044,2045]
    for _, row in timeline_df[timeline_df['year'].isin(key_years)].iterrows():
        print(f"{int(row['year']):<8} {row['annual_capex_billions']:<18.3f} {row['cumulative_capex_B']:<16.3f} {int(row['workers_retrained']):<20} {int(row['cumulative_workers'])}")

    total_capex   = plants_df['capex_billions'].sum()
    total_workers = plants_df['workers'].sum()
    total_retrain = plants_df['total_retrain_cost'].sum()

    print(f"\n{'='*55}")
    print(f"TRANSITION TOTALS V2")
    print(f"  Plants transitioning:       {len(plants_df)}")
    print(f"  Total capital investment:   ${total_capex:.2f}B")
    print(f"  Total workers affected:     {total_workers:,}")
    print(f"  Total retraining cost:      ${total_retrain:.1f}M")
    print(f"  Nuclear conversions:        {len(plants_df[plants_df['pathway']=='Nuclear_Conversion'])}")
    print(f"  Battery conversions:        {len(plants_df[plants_df['pathway']=='Battery_Storage'])}")
    print(f"{'='*55}")

    return plants_df, timeline_df

if __name__ == '__main__':
    gen_data, policy, demand_df = load_inputs()
    print("\n--- RUNNING V2 (feasibility-checked) ---")
    plants_df, timeline_df = run_transition_model_v2(gen_data, policy, demand_df)
