
import openpyxl, pandas as pd, sys, os
sys.path.insert(0, os.path.expanduser('~/Desktop/illinois_grid_model'))
from constants import *

def load_inputs(filepath=None):
    if filepath is None:
        filepath = os.path.expanduser('~/Desktop/illinois_grid_model/illinois_grid_model.xlsx')
    wb = openpyxl.load_workbook(filepath)
    ws_gen = wb['Inputs_Generation']
    gen_data = {}
    for row in ws_gen.iter_rows(min_row=2, values_only=True):
        if row[0]:
            gen_data[row[0]] = {'current_GW': row[1], 'capacity_factor': row[2],
                                'lcoe': row[3], 'capital_cost': row[4], 'lifetime': row[5]}
    ws_pol = wb['Inputs_Policy']
    policy = {}
    for row in ws_pol.iter_rows(min_row=2, values_only=True):
        if row[0]: policy[row[0]] = row[1]
    ws_dem = wb['Inputs_Demand']
    rows = []
    for row in ws_dem.iter_rows(min_row=2, values_only=True):
        if row[0] is not None:
            rows.append({'hour': row[0], 'weekday_MW': row[1],
                         'weekend_MW': row[2], 'peak_MW': row[3], 'season': row[4]})
    return gen_data, policy, pd.DataFrame(rows)

def run_storage_model_v2(gen_data, policy, demand_df, wind_GW=None, solar_GW=None):
    wind_GW  = wind_GW  or WIND_GW
    solar_GW = solar_GW or SOLAR_GW
    nuclear_MW = NUCLEAR_GW * NUCLEAR_CF * 1000
    solar_shape = [0,0,0,0,0,0.02,0.08,0.18,0.35,0.55,0.72,0.85,
                   0.92,0.95,0.93,0.85,0.70,0.50,0.28,0.10,0.03,0,0,0]
    wind_shape  = [0.72,0.75,0.78,0.80,0.82,0.80,0.72,0.60,0.50,0.42,0.38,0.35,
                   0.33,0.33,0.35,0.38,0.42,0.48,0.55,0.62,0.65,0.68,0.70,0.72]
    wind_season  = {'Summer_Peak':0.75,'Winter_Peak':1.15,'Spring_Low':1.10,'Fall_Low':0.95}
    solar_season = {'Summer_Peak':1.20,'Winter_Peak':0.55,'Spring_Low':0.90,'Fall_Low':0.80}
    results = []
    worst_deficit = 0
    for season in demand_df['season'].unique():
        s_df = demand_df[demand_df['season']==season].reset_index(drop=True)
        batt = 0; min_state = 0
        for _, row in s_df.iterrows():
            h = int(row['hour'])
            demand_MW = row['weekday_MW']
            wind_MW  = wind_GW  * 1000 * wind_shape[h]  * wind_season[season]
            solar_MW = solar_GW * 1000 * solar_shape[h] * solar_season[season]
            net = (nuclear_MW + wind_MW + solar_MW) - demand_MW
            batt += net; min_state = min(min_state, batt)
        worst_deficit = max(worst_deficit, abs(min_state))
        results.append({'season': season, 'deficit_MWh': abs(min_state),
                        'peak_MW': s_df['weekday_MW'].max()})
    results_df = pd.DataFrame(results)
    wind_e  = wind_GW  * WIND_CF
    solar_e = solar_GW * SOLAR_CF
    total_re = wind_e + solar_e
    if total_re > 0:
        balance = 1 - abs(wind_e - solar_e) / total_re
        comp_disc = WIND_SOLAR_STORAGE_DISCOUNT * balance
    else:
        balance = 0; comp_disc = 0
    raw_GWh  = worst_deficit / 1000
    disc_GWh = raw_GWh * (1 - comp_disc)
    cost_B   = disc_GWh * 1000 * (BATTERY_CAPEX_KW / 4) / 1e9
    print(f"\nSTORAGE MODEL V2  wind={wind_GW:.1f}GW  solar={solar_GW:.1f}GW")
    print(f"{'Season':<14} {'Peak MW':>10} {'Deficit MWh':>14}")
    print('-'*42)
    for _, r in results_df.iterrows():
        print(f"  {r['season']:<12} {r['peak_MW']:>10,.0f} {r['deficit_MWh']:>14,.0f}")
    print(f"  Worst deficit:             {raw_GWh:.1f} GWh (raw)")
    print(f"  Wind/solar balance:        {balance:.2f}  (1.0 = perfect 50/50)")
    print(f"  Complementarity discount:  {comp_disc*100:.1f}%")
    print(f"  Discounted storage need:   {disc_GWh:.1f} GWh")
    print(f"  Estimated cost:            ${cost_B:.2f}B")
    return disc_GWh, comp_disc

if __name__ == '__main__':
    gen_data, policy, demand_df = load_inputs()
    print("--- Current fleet ---")
    run_storage_model_v2(gen_data, policy, demand_df)
    print("--- 2045 target fleet (24GW wind, 12GW solar) ---")
    run_storage_model_v2(gen_data, policy, demand_df, wind_GW=24.0, solar_GW=12.0)
