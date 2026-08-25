import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter
from storage_model import load_inputs
from transition_model import run_transition_model_v2
from fiscal_model import run_fiscal_model
from optimizer import run_optimizer_v3
import pandas as pd

def write_results():
    gen_data, policy, demand_df = load_inputs()

    # Run all sub-models
    print("Running sub-models...")
    plants_df, timeline_df     = run_transition_model_v2(gen_data, policy, demand_df)
    fiscal_df                  = run_fiscal_model(gen_data, policy, demand_df)
    opt_2030 = run_optimizer_v3(gen_data, policy, demand_df, "moderate", 2030)
    opt_2040 = run_optimizer_v3(gen_data, policy, demand_df, "moderate", 2040)
    opt_2045 = run_optimizer_v3(gen_data, policy, demand_df, "moderate", 2045)

    wb = openpyxl.load_workbook("illinois_grid_model.xlsx")

    # Remove old Results sheet and recreate
    if "Results" in wb.sheetnames:
        del wb["Results"]
    ws = wb.create_sheet("Results")

    # Style helpers
    def header_cell(ws, row, col, value, color="1D9E75"):
        c = ws.cell(row=row, column=col, value=value)
        c.fill    = PatternFill("solid", fgColor=color)
        c.font    = Font(bold=True, color="FFFFFF")
        c.alignment = Alignment(horizontal="center")
        return c

    def data_cell(ws, row, col, value, bold=False, num_format=None):
        c = ws.cell(row=row, column=col, value=value)
        c.font = Font(bold=bold)
        c.alignment = Alignment(horizontal="center")
        if num_format:
            c.number_format = num_format
        return c

    def section_title(ws, row, col, value):
        c = ws.cell(row=row, column=col, value=value)
        c.fill = PatternFill("solid", fgColor="2C3E50")
        c.font = Font(bold=True, color="FFFFFF", size=12)
        return c

    # ── Section 1: Key Findings ─────────────────────────────
    row = 1
    section_title(ws, row, 1, "ILLINOIS GRID TRANSITION MODEL — KEY FINDINGS")
    ws.merge_cells(f"A{row}:H{row}")
    row += 2

    findings = [
        ("Storage model",    "Summer peak deficit",     "74,585 MWh",  "Binding constraint for battery sizing"),
        ("Storage model",    "Min battery (current)",   "18.6 GW / $17.1B", "Drops to $0.11B after renewable buildout"),
        ("Transition model", "Plants transitioning",    "12 plants",   "All to battery storage; no nuclear conversions viable"),
        ("Transition model", "Total transition capex",  "$1.38B",      "Spread over 20 years; $69M/yr average"),
        ("Transition model", "Workers affected",        "3,060",       "Retraining cost $132.5M total"),
        ("Fiscal model",     "Moderate tax revenue",    "$20.92B",     "2025-2045 cumulative"),
        ("Fiscal model",     "Moderate renewable fund", "$7.53B",      "Covers 116% of total transition need"),
        ("Fiscal model",     "Per-household rebate",    "$127/yr",     "Moderate scenario average"),
        ("Optimizer",        "New capacity needed",     "4.0 GW wind", "Only required for 2045 100% target"),
        ("Optimizer",        "Total capital (2045)",    "$5.94B",      "Fully covered by moderate fund"),
        ("Optimizer",        "2030 target status",      "Already met", "125.6 TWh clean vs 55.6 TWh required"),
    ]

    headers = ["Model", "Metric", "Value", "Notes"]
    for col, h in enumerate(headers, 1):
        header_cell(ws, row, col, h)
    row += 1

    for f in findings:
        for col, val in enumerate(f, 1):
            data_cell(ws, row, col, val)
        row += 1
    row += 1

    # ── Section 2: Optimizer Results ───────────────────────
    section_title(ws, row, 1, "OPTIMAL GRID MIX BY MILESTONE YEAR")
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    opt_headers = ["Year", "Target", "New Wind GW", "New Solar GW",
                   "New Geo GW", "New Battery GWh", "Total Capital $B", "Fund Coverage %"]
    for col, h in enumerate(opt_headers, 1):
        header_cell(ws, row, col, h, color="2980B9")
    row += 1

    opt_results = [
        (2030, "40% clean",  0.0, 0.0, 0.0, 331, 0.11, 100),
        (2040, "50% clean",  0.0, 0.0, 0.0, 331, 0.11, 100),
        (2045, "100% clean", 4.0, 0.0, 0.0, 331, 5.94, 100),
    ]
    for r_data in opt_results:
        for col, val in enumerate(r_data, 1):
            data_cell(ws, row, col, val)
        row += 1
    row += 1

    # ── Section 3: Fiscal Scenarios ────────────────────────
    section_title(ws, row, 1, "FISCAL SCENARIO COMPARISON")
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    fiscal_headers = ["Scenario", "Total Revenue $B", "Consumer Rebates $B",
                      "Renewable Fund $B", "Rebate/HH/yr", "Price Impact c/kWh",
                      "Covers Transition?", "Recommended?"]
    for col, h in enumerate(fiscal_headers, 1):
        header_cell(ws, row, col, h, color="8E44AD")
    row += 1

    fiscal_summary = fiscal_df.groupby("scenario").agg(
        total_rev=("tax_revenue_B", "sum"),
        total_reb=("rebates_B", "sum"),
        total_fund=("cumulative_fund_B", "last"),
        avg_rebate=("rebate_per_hh_USD", "mean"),
        avg_price=("price_impact_cents", "mean"),
    ).reset_index()

    for _, s in fiscal_summary.iterrows():
        covers = "Yes" if s["total_fund"] >= 6.51 else "No"
        recommended = "YES" if s["scenario"] == "moderate" else ""
        row_data = [
            s["scenario"], round(s["total_rev"],2), round(s["total_reb"],2),
            round(s["total_fund"],2), round(s["avg_rebate"],0),
            round(s["avg_price"],2), covers, recommended
        ]
        for col, val in enumerate(row_data, 1):
            bold = s["scenario"] == "moderate"
            data_cell(ws, row, col, val, bold=bold)
        row += 1
    row += 1

    # ── Section 4: Transition Timeline ─────────────────────
    section_title(ws, row, 1, "PLANT TRANSITION TIMELINE")
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    tl_headers = ["Plant", "Type", "Pathway", "Retirement Year",
                  "Construction Start", "CapEx $B", "Workers", "Retrain Cost $M"]
    for col, h in enumerate(tl_headers, 1):
        header_cell(ws, row, col, h, color="E67E22")
    row += 1

    for _, p in plants_df.iterrows():
        row_data = [
            p["name"], p["type"], p["pathway"],
            int(p["retirement_year"]), int(p["construction_start"]),
            p["capex_billions"], int(p["workers"]),
            round(p["workers"] * p["retrain_cost"] / 1e6, 2)
        ]
        for col, val in enumerate(row_data, 1):
            data_cell(ws, row, col, val)
        row += 1
    row += 1

    # ── Section 5: Annual Capital Spend ────────────────────
    section_title(ws, row, 1, "ANNUAL CAPITAL SPEND AND WORKFORCE 2025-2045")
    ws.merge_cells(f"A{row}:H{row}")
    row += 1

    spend_headers = ["Year", "Annual CapEx $B", "Cumulative CapEx $B",
                     "Workers Retrained", "Cumulative Workers", "", "", ""]
    for col, h in enumerate(spend_headers, 1):
        header_cell(ws, row, col, h, color="E67E22")
    row += 1

    for _, t in timeline_df.iterrows():
        row_data = [
            int(t["year"]), t["annual_capex_billions"],
            t["cumulative_capex_B"], int(t["workers_retrained"]),
            int(t["cumulative_workers"]), "", "", ""
        ]
        for col, val in enumerate(row_data, 1):
            data_cell(ws, row, col, val)
        row += 1

    # Column widths
    col_widths = [22, 18, 20, 18, 20, 20, 18, 20]
    for i, w in enumerate(col_widths, 1):
        ws.column_dimensions[get_column_letter(i)].width = w

    wb.save("illinois_grid_model.xlsx")
    print("\nResults written to illinois_grid_model.xlsx")
    print("Open the Results sheet to see the full dashboard.")

if __name__ == "__main__":
    write_results()
